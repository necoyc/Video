import matplotlib.pyplot as plt
import os
import numpy as np
import openpyxl as xl

EXCEL_BOOK_PATH = 'test1'

plt.rcParams['figure.subplot.left'] = 0.227
plt.rcParams['figure.subplot.right'] = 0.962
plt.rcParams['figure.subplot.bottom'] = 0.164
plt.rcParams['figure.subplot.top'] = 0.962
plt.rcParams['figure.subplot.wspace'] = 0.3
x_size = 3
y_size = 3


def save_plot(time_values: np.ndarray, y_values: np.ndarray, wrong_points: list, folder_name: str, filename: str, x_lable: str, y_label: str):
    """
    Сохраняет график в файл.
    
    :param time_values: Массив значений времени.
    :param y_values: Массив значений y.
    :param folder_name: Имя папки для сохранения файла.
    :param filename: Имя файла.
    :param x_lable: Подпись оси X.
    :param y_label: Подпись оси Y.
    """
    #добавить функцию для сглаживание данных
    y_values = np.delete(y_values, wrong_points)
    time_values = np.delete(time_values, wrong_points)
    plt.figure(figsize=(x_size, y_size))
    plt.plot(time_values, y_values)
    plt.xlabel(x_lable)
    plt.ylabel(y_label)
    plt.savefig(os.path.join(folder_name, f"{filename}.png"), dpi=300)
    plt.close()
def save_data_to_text(time_values: np.ndarray, y_values: np.ndarray, folder_name: str, filename: str) -> np.ndarray:
    """
    Сохраняет данные в текстовый файл и возвращает среднюю глубину.
    
    :param time_values: Массив значений времени.
    :param y_values: Многомерный массив значений y.
    :param folder_name: Имя папки для сохранения файла.
    :param filename: Имя файла.
    :return: Средняя глубина.
    """
    os.makedirs(folder_name, exist_ok=True)
    data_path = os.path.join(folder_name, f"{filename}.txt")
    
    average_depth = np.mean(y_values, axis=0)

    with open(data_path, "w") as file:
        for i in range(len(y_values[0])):
            column_values = [y[i] for y in y_values]
            file.write(f"{time_values[i]:.4f} {' '.join(map(str, column_values))}\n")

    wb = xl.Workbook()
    ws = wb.active
    
    wrong_points = []

    for i in range(len(time_values)):
        #if ((i < len(time_values) - 1) and abs(average_depth[i]) > abs(average_depth[i+1]*1.5)) or average_depth[i] < 0.0:
            #wrong_points.append(i)
            #continue
        ws.cell(row=i + 1, column=1, value=time_values[i])
        ws.cell(row=i + 1, column=2, value=str(average_depth[i]))

    os.makedirs(folder_name, exist_ok=True)
    wb.save(os.path.join(folder_name, filename + '.xlsx'))

    return average_depth, wrong_points
def process_data(time_values: np.ndarray, y_values: np.ndarray, folder_name: str, filename: str, x_lable: str, y_label: str):
    """
    Обрабатывает данные, сохраняет их в текстовый файл и строит график.
    
    :param time_values: Массив значений времени.
    :param y_values: Многомерный массив значений y.
    :param folder_name: Имя папки для сохранения файла.
    :param filename: Имя файла.
    :param x_lable: Подпись оси X.
    :param y_label: Подпись оси Y.
    """
    average_depth, wrong_points = save_data_to_text(time_values, y_values, folder_name, filename)
    save_plot(time_values, average_depth, wrong_points, folder_name, filename, x_lable, y_label)